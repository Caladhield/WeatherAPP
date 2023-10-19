import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import requests
from datetime import datetime

longitude = 18.02151508449004
latitude = 59.30996552541549

smhi_url = f'https://opendata-download-metfcst.smhi.se/api/category/pmp3g/version/2/geotype/point/lon/{longitude:.6f}/lat/{latitude:.6f}/data.json'

openweathermap_api_key = 'fa49fd29427dc75eb3b4febb8fa2c1b1'
openweathermap_url = f'https://api.openweathermap.org/data/3.0/onecall?lat=59.30&lon=18.02&exclude=minutely&appid={openweathermap_api_key}'

def get_smhi_data():
    response = requests.get(smhi_url)
    if response.status_code == 200:
        data = response.json()

        weather_data = []

        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S') #takes the datetime object and returns a string in the format we want

        for time_series in data['timeSeries']: # iterate through each dictionary
            timestamp = time_series['validTime'].replace('T', ' ').replace('Z', '') #access the validTime key in the timeSeries dictionary
            time = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S') # convert into a Python datetime object to later be able to calculate local time

            local_time = (time - datetime.now()).total_seconds() / 3600

            if 0 <= local_time <= 24:
                date = timestamp.split()[0]
                hour = int(timestamp.split()[1][:2]) # splits the timestamp at the space and selects the second part, which represents the time. It then extracts the hour

                temperature = 0
                precipitation = 0
                provider = 'SMHI'

                for items in time_series['parameters']:
                    if items['name'] == 't':
                        temperature = items['values'][0] # 0 because it only has 1 item in the list and thats the value we want
                    elif items['name'] == 'pcat':
                        precipitation = items['values'][0]
                
                if precipitation > 0: 
                    precipitation = "Precipitation"
                else:
                    precipitation ="No precipitation"

                weather_data.append({
                    'Created': current_time,
                    'Longitude': longitude,
                    'Latitude': latitude,
                    'Date': date,
                    'Hour': hour,
                    'Temperature (°C)': temperature,
                    'Precipitation': precipitation,
                    'Provider': provider,
                })

        dataframe = pd.DataFrame(weather_data)
        
        file_path = "Weather_data_SMHI.xlsx"

        excel = pd.ExcelWriter(file_path, engine='openpyxl')

        dataframe.to_excel(excel, index=False, sheet_name='SmhiData')

        workbook = excel.book
        worksheet = excel.sheets['SmhiData']

        column_widths = {'A': 20, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 20, 'G': 20, 'H': 15}

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

        workbook.save(file_path)
        print(f'SMHI Data saved to {file_path}')

    else:
        print(f'SMHI Data Error: {response.status_code}')

def get_openweathermap_data():
    response = requests.get(openweathermap_url)       
    if response.status_code == 200:
        data = response.json()
        
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        weather_data = []
        
        now = datetime.now()
        next_24_hours = now + pd.DateOffset(hours=24)
        time_in_24_hours = int(next_24_hours.timestamp())

        for item in data['hourly']:
            time = item['dt']
            if time > now.timestamp() and time <= time_in_24_hours:
                temp_kelvin = item['temp']
                temp_celsius = temp_kelvin - 273.15
                weather = item['weather'][0]['main']
                hour = datetime.utcfromtimestamp(time).strftime('%H')
                local_time = int(hour)+1
                date = datetime.utcfromtimestamp(time).strftime('%Y-%m-%d')

                if weather == 'Rain':
                    precipitation = 'Rain'
                elif weather == 'Snow':
                    precipitation = 'Snow'
                elif weather == 'Drizzle':
                    precipitation == 'Drizzle'
                else:
                    precipitation = 'No precipitation'

                weather_data.append({
                    'Created': current_time,
                    'Longitude': longitude,
                    'Latitude': latitude,
                    'Date': date,
                    'Hour': local_time,
                    'Temperature (°C)': temp_celsius,
                    'Precipitation': precipitation,
                    'Provider': 'OWM',
                    
                })

        weather_dataframe = pd.DataFrame(weather_data)

        file_path = "Weather_Data_OpenWeatherMap.xlsx"

        excel = pd.ExcelWriter(file_path, engine='openpyxl')

        weather_dataframe.to_excel(excel, index=False, sheet_name='OWMData')

        workbook = excel.book
        worksheet = excel.sheets['OWMData']

        column_widths = {'A': 20, 'B': 20, 'C': 20, 'D': 20, 'E': 20, 'F': 20, 'G': 20, 'H': 20}

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

        workbook.save(file_path)
        print(f'OpenWeatherMap Data saved to {file_path}')

    else:
        print(f'OpenWeatherMap Error: {response.status_code}')

def combine_and_save_data():
    smhi_data = pd.read_excel("Weather_data_SMHI.xlsx")
    owm_data = pd.read_excel("Weather_Data_OpenWeatherMap.xlsx")
    combined_data = pd.concat([smhi_data, owm_data])

    file_path = "Combined_Weather_Data.xlsx"

    excel = pd.ExcelWriter(file_path, engine='openpyxl')

    combined_data.to_excel(excel, index=False, sheet_name='WeatherData')
    workbook = excel.book
    worksheet = excel.sheets['WeatherData']

    column_widths = {'A': 20, 'B': 20, 'C': 20, 'D': 20, 'E': 20, 'F': 20, 'G': 20, 'H': 20}

    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    workbook.save(file_path)
    print(f'Combined Data saved to {file_path}')

def print_forecast(filename):
    try:
        dataframe = pd.read_excel(filename)
        print(dataframe)
    except FileNotFoundError:
        print(f'The file {filename} does not exist.')

while True:
    print("Menu:")
    print("1. Get SMHI Data")
    print("2. Get OpenWeatherMap Data")
    print("3. Combine Data")
    print("4. Print SMHI prognos.")
    print("5. Print OWM prognos.")
    print("6. Print Forecast")
    print("7. Exit")

    choice = input("Select an option: ")
    if choice == '1':
        get_smhi_data()
    elif choice == '2':
        get_openweathermap_data()
    elif choice == '3':
        combine_and_save_data()
    elif choice == '4':
        print_forecast("Weather_data_SMHI.xlsx")
    elif choice == '5':
        print_forecast("Weather_data_OpenWeatherMap.xlsx")
    elif choice == '6':
        print_forecast("Combined_Weather_Data.xlsx")
    elif choice == '7':
        print("Exiting the program.")
        break
    else:
        print("Please choose a valid option (1, 2, 3, 4, or 5).")
